import xlsxwriter  # For creating new Excel file if it doesn't exist
from openpyxl import load_workbook

def main():
    print("Welcome to ABC Bank")
    choice = input("Press 1 for Login \nPress 2 for Sign in\n:")
    match choice:
        case "1":
            username = input("Enter your Username : ").casefold().strip().replace(" ", "")
            password = input("Enter your Password : ").strip()
            exe_data = load_excel_data('D:/Project_py/Data.xlsx', 'Sheet1')
            
            # Check if username and password match
            for row in exe_data:
                if username == row[0] and password == row[1]:
                    print("Login successful")
                    while True:
                        acc_oper = input("1. Cash Deposit \n2. Cash Withdrawal \n3. Balance Inquiry \n4. Logout : ")
                        match acc_oper:
                            case "1":
                                amount = float(input("Enter amount to deposit: "))
                                update_balance(username, amount, 'credit')
                                print("Cash Deposit successful")
                            case "2":
                                amount = float(input("Enter amount to withdraw: "))
                                update_balance(username, -amount, 'withdraw')
                                print("Cash Withdrawal successful")
                            case "3":
                                balance = get_balance(username)
                                print(f"Your current balance is: {balance}")
                            case "4":
                                print("Logging out...")
                                return
                            case _:
                                print("Invalid option, please try again.")
                    break
            else:
                print('Login failed')

        case "2":
            print("New Account")
            first_name = input("Enter First Name : ").title()
            last_name = input("Enter Last Name : ").title()
            email = input("Enter Email Address : ")
            acc_name = (first_name + last_name).casefold().strip().replace(" ", "")
            while True:
                pass_new_acc = input("Set password : ").strip()
                re_new_acc = input("Re-enter password : ").strip()
                if pass_new_acc == re_new_acc:
                    print("Password set successfully")
                    new_data = [acc_name, pass_new_acc, 0.0, 0.0, 0.0]  # Initial values for balance, credit, withdrawal
                    add_data_to_new_row('D:/Project_py/Data.xlsx', new_data)
                    break
                else:
                    print("Passwords do not match, please try again.")

def load_excel_data(file_path, sheet_name):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]

        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return []

def add_data_to_new_row(file_path, new_data):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Find the next available row
        next_row = sheet.max_row + 1

        # Write the new data to the next row
        for col, value in enumerate(new_data, start=1):
            sheet.cell(row=next_row, column=col, value=value)

        workbook.save(file_path)
        print(f"Data added successfully to row {next_row}")

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        # Create a new file with xlsxwriter and add the data
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet()
        headers = ["Username", "Password", "Balance", "Credit", "Withdrawal"]
        worksheet.write_row(0, 0, headers)  # Add headers
        worksheet.write_row(1, 0, new_data)  # Add the new data
        workbook.close()
        print(f"New file created and data saved to: {file_path}")

def update_balance(username, amount, operation):
    file_path = 'D:/Project_py/Data.xlsx'
    data = load_excel_data(file_path, 'Sheet1')
    for i, row in enumerate(data):
        if username == row[0]:
            balance, credit, withdrawal = row[2], row[3], row[4]
            if operation == 'credit':
                new_balance = balance + amount
                new_credit = credit + amount
                new_withdrawal = withdrawal
            elif operation == 'withdraw':
                new_balance = balance - amount
                new_credit = credit
                new_withdrawal = withdrawal - amount
            else:
                return
            data[i] = [username, row[1], new_balance, new_credit, new_withdrawal]
            save_excel_data(file_path, 'Sheet1', data)
            break

def get_balance(username):
    file_path = 'D:/Project_py/Data.xlsx'
    data = load_excel_data(file_path, 'Sheet1')
    for row in data:
        if username == row[0]:
            return row[2]
    return 0.0

def save_excel_data(file_path, sheet_name, data):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    for i, row in enumerate(data, start=1):
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=value)

    workbook.save(file_path)
main()
